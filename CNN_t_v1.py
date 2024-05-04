# -*- coding: utf-8 -*-
"""
Created on Tue Jul 11 15:38:00 2017

@author: MinaMelek
"""
from __future__ import absolute_import
import tensorflow as tf
import numpy as np
from openpyxl import load_workbook,Workbook
import time
import ReadTrain
from tensorflow.python.tools import freeze_graph, optimize_for_inference_lib
print("Initializing...")
height = 40#int(292*3/4)
width = 80#int(548*3/4)
channels = 3
classes_num = 7
image_size = height*width*channels
LR = 0.001 # Learning rate
batch_size = 32 # batch size
x = tf.placeholder(tf.float32, [None, image_size], name='IN')
y = tf.placeholder(tf.float32, [None, classes_num])
keep_prob1 = tf.placeholder(tf.float32)
keep_prob2 = tf.placeholder(tf.float32)
keep_prob3 = tf.placeholder(tf.float32)
keep_probf = tf.placeholder(tf.float32)
dirct = "./model.ckpt"
dir_w = "./modelWeights.ckpt"
dir_b = "./modelBias.ckpt"
sess = tf.Session() # tf.InteractiveSession()


def weight_variable(shape):
  initial = tf.truncated_normal(shape, stddev=0.1)
  return tf.Variable(initial)

def bias_variable(shape):
  initial = tf.constant(0.1, shape=shape)
  return tf.Variable(initial)
  
def conv2d(x, W, name=None):
  return tf.nn.conv2d(x, W, strides=[1, 1, 1, 1], padding='SAME', name=name)

def max_pool_2x2(x, name=None):
  return tf.nn.max_pool(x, ksize=[1, 2, 2, 1],
                        strides=[1, 2, 2, 1], padding='SAME',name=name)

def extend_labels(labels,classes):
    sets_num = labels.shape[0]
    ex = np.reshape([False]* sets_num *classes,[sets_num,classes])
    for n in range(sets_num): ex[n,int(labels[n])] = True
    return ex
def ConvolutionalNeuralNetwork(x):#, keep_rate):
    weights = {
        # 3 x 3 convolution, 1 input image, 32 outputs
        'W_conv11': tf.Variable(tf.random_normal([3, 3, channels, 64],stddev=0.01),name='conv11_w'),
        # 3 x 3 conv, 32 inputs, 64 outputs
        'W_conv12': tf.Variable(tf.random_normal([3, 3, 64, 128],stddev=0.01),name='conv12_w'),
        # 3 x 3 convolution, 1 input image, 32 outputs
        'W_conv21': tf.Variable(tf.random_normal([3, 3, 128, 256],stddev=0.01),name='conv21_w'),
        # 3 x 3 conv, 32 inputs, 64 outputs
        'W_conv22': tf.Variable(tf.random_normal([3, 3, 256, 256],stddev=0.01),name='conv22_w'),
        # 3 x 3 convolution, 1 input image, 32 outputs
        'W_conv31': tf.Variable(tf.random_normal([3, 3, 256, 256],stddev=0.01),name='conv31_w'),
        # 3 x 3 conv, 32 inputs, 64 outputs
        'W_conv32': tf.Variable(tf.random_normal([3, 3, 256, 512],stddev=0.01),name='conv32_w'),
        # fully connected, 8*8*64 inputs, 1024 outputs
        'W_fc': tf.Variable(tf.random_normal([5*10*512, 1024],stddev=0.01),name='fc1_w'),
        # 1024 inputs, 10 outputs (class prediction)
        'out': tf.Variable(tf.random_normal([1024, classes_num],stddev=0.01),name='out_w')
    }
    
    biases = {
        'b_conv11': tf.Variable(tf.ones([64]),name='conv11_b'),
        'b_conv12': tf.Variable(tf.ones([128]),name='conv12_b'),
        'b_conv21': tf.Variable(tf.ones([256]),name='conv21_b'),
        'b_conv22': tf.Variable(tf.ones([256]),name='conv22_b'),
        'b_conv31': tf.Variable(tf.ones([256]),name='conv31_b'),
        'b_conv32': tf.Variable(tf.ones([512]),name='conv32_b'),
        'b_fc': tf.Variable(tf.ones([1024]),name='fc1_b'),
        'out': tf.Variable(tf.ones([classes_num]),name='out_b')
    }
    
    global saver_w
    global saver_b
    saver_w = tf.train.Saver(weights)
    saver_b = tf.train.Saver(biases)
    
    
    # Reshape input to a 4D tensor
    x = tf.reshape(x, shape=[-1, height , width , channels ])
    
    
    # First Convolution Layer, using our function
    conv1 = tf.nn.relu(conv2d(x, weights['W_conv11'],name='CONV11') + biases['b_conv11'])
    # Convolution Layer, using our function
    conv1 = tf.nn.relu(conv2d(conv1, weights['W_conv12'],name='CONV12') + biases['b_conv12'])
    # Max Pooling (down-sampling)
    conv1 = max_pool_2x2(conv1, name='MAX_POOL1')
    # Dropout
    conv1_drop = tf.nn.dropout(conv1, keep_prob1, name='DROPOUT1')
    # Second Convolution Layer, using our function
    conv2 = tf.nn.relu(conv2d(conv1_drop, weights['W_conv21'],name='CONV21') + biases['b_conv21'])
    # Convolution Layer, using our function
    conv2 = tf.nn.relu(conv2d(conv2, weights['W_conv22'],name='CONV22') + biases['b_conv22'])
    # Max Pooling (down-sampling)
    conv2 = max_pool_2x2(conv2, name='MAX_POOL2')
    # Dropout
    conv2_drop = tf.nn.dropout(conv2, keep_prob2, name='DROPOUT2')   
    # Third Convolution Layer, using our function
    conv3 = tf.nn.relu(conv2d(conv2_drop, weights['W_conv31'],name='CONV31') + biases['b_conv31'])
    # Convolution Layer, using our function
    conv3 = tf.nn.relu(conv2d(conv3, weights['W_conv32'],name='CONV32') + biases['b_conv32'])
    # Max Pooling (down-sampling)
    conv3 = max_pool_2x2(conv3, name='MAX_POOL3')
    # Dropout
    conv3_drop = tf.nn.dropout(conv3, keep_prob3, name='DROPOUT3')    
    # Fully connected layer
    # Reshape conv2 output to fit fully connected layer
    fc = tf.reshape(conv3_drop, [-1, 5*10*512])
    fc = tf.nn.relu(tf.matmul(fc, weights['W_fc'],name='FC1') + biases['b_fc'])
    # Dropout
    fc = tf.nn.dropout(fc, keep_probf, name='DROPOUTf')
    output = tf.matmul(fc, weights['out'], name='OUT') + biases['out']
    return output

#def buildConvolutionLayer(x,l_in,l1,l2):
#    W_conv11 = weight_variable([3, 3, l_in, l1])
#    b_conv11 = bias_variable([l1])
#    W_conv12 = weight_variable([3, 3, l1, l2])
#    b_conv12 = bias_variable([l2])
#    # applying 2 convolution layers and a subsampling layer
#    h_conv1 = tf.nn.relu(conv2d(x, W_conv11) + b_conv11)
#    h_conv2 = tf.nn.relu(conv2d(h_conv1, W_conv12) + b_conv12)
#    h_pool = max_pool_2x2(h_conv2)
#    return h_pool
######################################################
print('Reading Data....')
loaded = np.load('/home/ti1080/GP_Team#/drive-download-20170709T124015Z-001/Trian_data.npz')
y1 = loaded['y_train']
x_train = loaded['x_train']
loaded = np.load('/home/ti1080/GP_Team#/drive-download-20170709T124015Z-001/Val_data.npz')
y2 = loaded['y_val']
x_val = loaded['x_val']
loaded = np.load('/home/ti1080/GP_Team#/drive-download-20170709T124015Z-001/Test_data.npz')
y3 = loaded['y_test']
x_test = loaded['x_test']
del loaded
y_train = extend_labels(y1,classes_num)
y_val = extend_labels(y2,classes_num)
y_test = extend_labels(y3,classes_num)

# checks
print('train shape:','\nData: ',x_train.shape,'\nlabels: ',y_train.shape)
print('val shape:','\nData: ',x_val.shape,'\nlabels: ',y_val.shape)
print('test shape:','\nData: ',x_test.shape,'\nlabels: ',y_test.shape)

### Convert to 4D tensor
#x_train = tf.reshape(x_train, shape=x_train.shape)
#x_val = tf.reshape(x_val, shape=x_val.shape)
#x_test = tf.reshape(x_test, shape=x_test.shape)

train = ReadTrain.Organize(x_train, y_train)
validation = ReadTrain.Organize(x_val, y_val)
test = ReadTrain.Organize(x_test,y_test)
#layers=[[32,64,64,128,256],[64,64,64,64,64],[128,128,256,256,512],[32,32,64,64,128]
#,[64,64,128,128,256],[64,128,128,256,512],[32,64,128,256,0],[64,128,256,256,512],[64,128,128,256,256]]
#l=7 # This variable takes numbers from 0 to 8 to refers to the index of the previouse list more detailes in APPENDIX III
#print("at L",l)
## First Convolutional Layers
#W_conv11 = weight_variable([3, 3, channels, layers[l][0]])
#b_conv11 = bias_variable([layers[l][0]])
#W_conv12 = weight_variable([3, 3, layers[l][0], layers[l][1]])
#b_conv12 = bias_variable([layers[l][1]])
## applying 2 convolution layers and a subsampling layer
#h_conv11 = tf.nn.relu(conv2d(x_train, W_conv11) + b_conv11)
#h_conv12 = tf.nn.relu(conv2d(h_conv11, W_conv12) + b_conv12)
#h_pool1 = max_pool_2x2(h_conv12)
##h_pool1 = buildConvolutionLayer(x_train, channels, layers[l][0], layers[l][1])
## Dropout
#keep_prob1 = tf.placeholder(tf.float32)
#h_pool1_drop = tf.nn.dropout(h_pool1, keep_prob1)
##******************************************************************************
## Second Convolutional Layers
#W_conv21 = weight_variable([3, 3, layers[l][1], layers[l][2]])
#b_conv21 = bias_variable([layers[l][2]])
#W_conv22 = weight_variable([3, 3, layers[l][2], layers[l][3]])
#b_conv22 = bias_variable([layers[l][3]])
## applying 2 convolution layers and a subsampling layer
#h_conv21 = tf.nn.relu(conv2d(h_pool1_drop, W_conv21) + b_conv21)
#h_conv22 = tf.nn.relu(conv2d(h_conv21, W_conv22) + b_conv22)
#h_pool2 = max_pool_2x2(h_conv22)
##h_pool2 = buildConvolutionLayer(h_pool1_drop, layers[l][1], layers[l][2], layers[l][3])
## Dropout
#keep_prob2 = tf.placeholder(tf.float32)
#h_pool2_drop = tf.nn.dropout(h_pool2, keep_prob2)
##******************************************************************************
## Densely Connected Layer
#W_fc1 = weight_variable([10 * 20 * layers[l][3], 256])
#b_fc1 = bias_variable([256])
#
#h_pool2_flat = tf.reshape(h_pool2, [-1, 10*20*layers[l][3]])
#h_fc1 = tf.nn.relu(tf.matmul(h_pool2_flat, W_fc1) + b_fc1)
## Dropout
#keep_prob3 = tf.placeholder(tf.float32)
#h_fc1_drop = tf.nn.dropout(h_fc1, keep_prob3)
## Readout Layer
#W_fc2 = weight_variable([256, 7])
#b_fc2 = bias_variable([7])
#
#y_conv = tf.matmul(h_fc1_drop, W_fc2) + b_fc2
#******************************************************************************

y_conv = ConvolutionalNeuralNetwork(x)
saver1 = tf.train.Saver()
# Training
#y_pred_cls = tf.argmax(y_pred, dimension=1)
checkpoint = 0
cost = tf.reduce_mean(
    tf.nn.softmax_cross_entropy_with_logits(labels=y, logits=y_conv))
train_step = tf.train.AdamOptimizer(learning_rate=LR).minimize(cost)#, epsilon=1e-6
correct_prediction = tf.equal(tf.argmax(y_conv, 1), tf.argmax(y, 1))
accuracy = tf.reduce_mean(tf.cast(correct_prediction, tf.float32))
epochs = 5
with tf.Session() as sess:
    sess.run(tf.global_variables_initializer())
    # save the graph
    tf.train.write_graph(sess.graph_def, '.', 'tfdroid.pbtxt') 
    tf.summary.FileWriter('./graphs',sess.graph)
    saver = tf.train.Saver()
    if not(checkpoint==0):  
        saver.restore(sess, dirct)
        book = load_workbook("model.xlsx")
        sheet1 = book.get_sheet_by_name("Sheet2")
    else :
        book = Workbook()
        sheet1 = book.create_sheet("Sheet2") #.add_sheet('Sheet1',cell_overwrite_ok=True)
        book.save('model.xlsx')
    for epoch in range(epochs):     
        epoch_loss = 0
        start_time = time.time()
        for _ in range(int(train.num_examples/batch_size)):
            epoch_x, epoch_y = train.next_batch(batch_size) # TODO: change to train.next_batch
            _, c = sess.run([train_step, cost], feed_dict={x: epoch_x, y: epoch_y, keep_prob1: 0.75, keep_prob2: 0.75, keep_prob3: 0.75, keep_probf: 0.5})
            epoch_loss += c
        
        duration = time.time() - start_time
        print("Epoch lasts for : ",duration)
        print('Epoch', epoch, 'completed out of',epochs,'loss:',epoch_loss)
        # save a checkpoint
        saver.save(sess,dirct)
        saver_w.save(sess,dir_w)
        saver_b.save(sess,dir_b)
        
        checkpoint = 1
        print("weights have been saved")
        # test against the validation data set
        percent = 0
        for batch in range(int(validation.num_examples/batch_size)):
            val_x, val_y = validation.next_batch(batch_size)
            percent += accuracy.eval({x:val_x, y:val_y, keep_prob1: 1, keep_prob2: 1, keep_prob3: 1, keep_probf: 1})
            #print percent
        percent /= (validation.num_examples/batch_size)
        print('Validation Accuracy:', percent)
        sheet1.append({"A":percent})
        book.save('model.xlsx')    
        
        if(epoch%10==9 or epoch==epochs-1):
            # check accuracy aginst test data
            percent = 0
            ##################################################
            for batch in range(int(test.num_examples/1000)):
                test_x, test_y = test.next_batch(1000) 
                percent += accuracy.eval({x:test_x, y:test_y, keep_prob1: 1, keep_prob2: 1, keep_prob3: 1, keep_probf: 1})
            print('Test Accuracy:',percent/10)
    #save a checkpoint file, which will store the above assignment  
    saver1.save(sess, 'tfdroid.ckpt')
# Freexing Graph
MODEL_NAME = 'tfdroid'

# Freeze the graph

input_graph_path = MODEL_NAME+'.pbtxt'
checkpoint_path = './'+MODEL_NAME+'.ckpt'
input_saver_def_path = ""
input_binary = False
output_node_names = "OUT"
restore_op_name = "save/restore_all"
filename_tensor_name = "save/Const:0"
output_frozen_graph_name = 'frozen_'+MODEL_NAME+'.pb'
output_optimized_graph_name = 'optimized_'+MODEL_NAME+'.pb'
clear_devices = True


freeze_graph.freeze_graph(input_graph_path, input_saver_def_path,
                          input_binary, checkpoint_path, output_node_names,
                          restore_op_name, filename_tensor_name,
                          output_frozen_graph_name, clear_devices, "")
# 

input_graph_def = tf.GraphDef()
with tf.gfile.Open(output_frozen_graph_name, "r") as f:
    data = f.read()
    input_graph_def.ParseFromString(data)

output_graph_def = optimize_for_inference_lib.optimize_for_inference(
        input_graph_def,
        ["IN"], # an array of the input node(s)
        ["OUT"], # an array of output nodes
        tf.float32.as_datatype_enum)

# Save the optimized graph

f = tf.gfile.FastGFile(output_optimized_graph_name, "w")
f.write(output_graph_def.SerializeToString())
