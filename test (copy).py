#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Sat May 20 21:59:23 2017

@author: salama
"""

from __future__ import absolute_import
import tensorflow as tf
import numpy as np
import readCIFAR
import time
#import xlwt
#import openpyxl
from openpyxl import load_workbook,Workbook

Xtr, Ytr, Xts, Yts = readCIFAR.load_CIFAR10("cifar-10-batches-py")

def make_one_hot(labels,num_classes):
    num_examples = labels.shape[0]
    one_hot = np.reshape([False]* num_examples *num_classes,[num_examples,num_classes])
    one_hot[range(num_examples),labels] = True
    return one_hot

def cross_validation(Xtr,Ytr,num_folds,epoch):
    fold_size = Xtr.shape[0] / num_folds
    val_part = epoch % num_folds
    temp = np.reshape(range(Xtr.shape[0]),(-1))
    mask1 =  temp < ((val_part+1)*fold_size)
    mask2 = temp >= (val_part*fold_size)
    #mask[val_part*fold_size:(val_part+1)*fold_size] = True
    mask = mask1 * mask2
    x_val = Xtr[mask]
    y_val = Ytr[mask]
    x_train = Xtr[~mask]
    y_train = Ytr[~mask]
    return x_train, y_train, x_val, y_val


batch_size = 200
n_classes = 10 # CIFAR10 total classes (0-9 objects)
image_height = 32 
image_width =  32
image_channels = 3
image_size = image_height * image_width *image_channels
learning_rate = 0.001

"""
mask = np.random.choice(range(50000),100, replace=False)
X_dev = Xtr[mask]
y_dev = make_one_hot(Ytr[mask],n_classes)
develop = readCIFAR.DataSet(X_dev,y_dev)




train = readCIFAR.DataSet(Xtr[:49000],train_labels)
validation = readCIFAR.DataSet(Xtr[49000:],val_labels)
"""
test_labels = make_one_hot(Yts,n_classes)
test = readCIFAR.DataSet(Xts,test_labels)


# tf Graph input
x = tf.placeholder(tf.float32, [None, image_size])
y = tf.placeholder(tf.float32, [None, n_classes])
#y_test = tf.placeholder(tf.int32,[None,n_classes])
keep_prob = tf.placeholder(tf.float32)


#file_path = '/model.ckpt'

def conv2d(x, W, name=None):
  return tf.nn.conv2d(x, W, strides=[1, 1, 1, 1], padding='SAME', name=name)

def maxpool2d(x, name=None):
  return tf.nn.max_pool(x, ksize=[1, 2, 2, 1], strides=[1, 2, 2, 1], padding='SAME', name=name)
  
#def convolutional_neural_network(x):#, keep_rate):
#    weights = {
#        # 5 x 5 convolution, 1 input image, 32 outputs
#        'W_conv1': tf.Variable(tf.random_normal([3,3, image_channels , 32])),
#        #'W_conv11': tf.Variable(tf.random_normal([3, 3, 32 , 32])),
#        # 5x5 conv, 32 inputs, 64 outputs 
#        'W_conv2': tf.Variable(tf.random_normal([3, 3,32, 64])),
#        #'W_conv22': tf.Variable(tf.random_normal([3, 3, 64, 64])),
#        # fully connected, 8*8*64 inputs, 1024 outputs
#        'W_fc': tf.Variable(tf.random_normal([8*8*64, 1024])),
#        'W_fc_2': tf.Variable(tf.random_normal([1024, 256])),
#        # 1024 inputs, 10 outputs (class prediction)
#        'out': tf.Variable(tf.random_normal([256, n_classes]))
#    }
#    # add weight decay
#    wd = 0.004
#    #tf.add_to_collection('losses',tf.multiply(tf.nn.l2_loss(weights['W_conv1']),wd))
#    #tf.add_to_collection('losses',tf.multiply(tf.nn.l2_loss(weights['W_conv11']),wd))
#    #tf.add_to_collection('losses',tf.multiply(tf.nn.l2_loss(weights['W_conv2']),wd))
#    #tf.add_to_collection('losses',tf.multiply(tf.nn.l2_loss(weights['W_conv22']),wd))
#    tf.add_to_collection('losses',tf.multiply(tf.nn.l2_loss(weights['W_fc']),wd))
#    tf.add_to_collection('losses',tf.multiply(tf.nn.l2_loss(weights['W_fc_2']),wd))
#    #tf.add_to_collection('losses',tf.multiply(tf.nn.l2_loss(weights['out']),wd))
#    
#    
#    biases = {
#        'b_conv1': tf.Variable(tf.constant(0.0,shape=[32])),
#        #'b_conv11': tf.Variable(tf.random_normal([32])),
#        'b_conv2': tf.Variable(tf.constant(0.0,shape=[64])),
#        #'b_conv22': tf.Variable(tf.random_normal([64])),
#        'b_fc': tf.Variable(tf.constant(0.0,shape=[1024])),
#        'b_fc_2': tf.Variable(tf.constant(0.0,shape=[256])),
#        'out': tf.Variable(tf.constant(0.0,shape=[n_classes]))
#    }
#     # Reshape input to a 4D tensor 
#    x = tf.reshape(x, shape=[-1, image_height , image_width , image_channels ])
#    
#    
#    # Convolution Layer, using our function
#    conv1 = tf.nn.relu(conv2d(x, weights['W_conv1']) + biases['b_conv1'])
#    #conv11 = tf.nn.relu(conv2d(conv1, weights['W_conv11']) + biases['b_conv11'])
#    #norm1 = tf.nn.lrn(conv11, 4, bias=1.0, alpha=0.001 / 9.0, beta=0.75,
#    #                name='norm1')
#    #conv1 = tf.nn.dropout(conv1, keep_prob)
#    # Max Pooling (down-sampling)
#    max1 = maxpool2d(conv1)
#    
#    
#    # Convolution Layer
#    conv2 = tf.nn.relu(conv2d(max1, weights['W_conv2']) + biases['b_conv2'])
#    #conv22 = tf.nn.relu(conv2d(conv2, weights['W_conv22']) + biases['b_conv22'])
#    #norm2 = tf.nn.lrn(conv22, 4, bias=1.0, alpha=0.001 / 9.0, beta=0.75,
#    #                name='norm2')
#    #conv2 = tf.nn.dropout(conv2, keep_prob)
#    # Max Pooling (down-sampling)
#    max2 = maxpool2d(conv2)
#    
#    
#    # Fully connected layer
#    # Reshape conv2 output to fit fully connected layer
#    fc = tf.reshape(max2, [-1, 8*8*64])
#    fc_r = tf.nn.relu(tf.matmul(fc, weights['W_fc']) + biases['b_fc'])
#    #fc = tf.nn.dropout(fc, keep_prob)
#    
#    fc_2 = tf.nn.relu(tf.matmul(fc_r, weights['W_fc_2']) + biases['b_fc_2'])
#    #fc_2 = tf.nn.dropout(fc_2, keep_prob)
#
#
#    output = tf.matmul(fc_2, weights['out']) + biases['out']
#    
#    #print weights['W_conv1'],weights['W_conv22'],weights['W_fc_2']
#    return output


def convolutional_neural_network(x):#, keep_rate):
    weights = {
        # 5 x 5 convolution, 1 input image, 32 outputs
        'W_conv1': tf.Variable(tf.random_normal([5, 5, image_channels , 32],stddev=0.01),name='conv1_w'),
        'W_conv1_2': tf.Variable(tf.random_normal([5, 5, 32 , 32],stddev=0.01),name='conv1_2_w'),
        # 5x5 conv, 32 inputs, 64 outputs 
        'W_conv2': tf.Variable(tf.random_normal([5, 5, 32, 64],stddev=0.01),name='conv2_w'),
        'W_conv2_2': tf.Variable(tf.random_normal([5, 5, 64, 64],stddev=0.01),name='conv2_2_w'),
        # fully connected, 8*8*64 inputs, 1024 outputs
        'W_fc': tf.Variable(tf.random_normal([8*8*64, 1024],stddev=0.01),name='fc1_w'),
        # 1024 inputs, 10 outputs (class prediction)
        'out': tf.Variable(tf.random_normal([1024, n_classes],stddev=0.01),name='out_w')
    }
    
    biases = {
        'b_conv1': tf.Variable(tf.ones([32]),name='conv1_b'),
        'b_conv1_2': tf.Variable(tf.ones([32]),name='conv1_2_b'),
        'b_conv2': tf.Variable(tf.ones([64]),name='conv2_b'),
        'b_conv2_2': tf.Variable(tf.ones([64]),name='conv2_2_b'),
        'b_fc': tf.Variable(tf.ones([1024]),name='fc1_b'),
        'out': tf.Variable(tf.ones([n_classes]),name='out_b')
    }
    
    global saver_w
    global saver_b
    saver_w = tf.train.Saver(weights)
    saver_b = tf.train.Saver(biases)
    
    
    # Reshape input to a 4D tensor 
    x = tf.reshape(x, shape=[-1, image_height , image_width , image_channels ])
    
    
    # Convolution Layer, using our function
    conv1 = tf.nn.relu(conv2d(x, weights['W_conv1'],name='CONV1') + biases['b_conv1'])
    conv1 = tf.nn.relu(conv2d(conv1, weights['W_conv1_2'],name='CONV1_2') + biases['b_conv1_2'])
    # Max Pooling (down-sampling)
    conv1 = maxpool2d(conv1, name='MAX_POOL1')
    
    
    # Convolution Layer
    conv2 = tf.nn.relu(conv2d(conv1, weights['W_conv2'],name='CONV2') + biases['b_conv2'])
    conv2 = tf.nn.relu(conv2d(conv2, weights['W_conv2_2'],name='CONV2_2') + biases['b_conv2_2'])
    # Max Pooling (down-sampling)
    conv2 = maxpool2d(conv2, name='MAX_POOL2')
    
    
    # Fully connected layer
    # Reshape conv2 output to fit fully connected layer
    fc = tf.reshape(conv2, [-1, 8*8*64])
    fc = tf.nn.relu(tf.matmul(fc, weights['W_fc'],name='FC1') + biases['b_fc'])
    
    fc = tf.nn.dropout(fc, keep_prob, name='DROPOUT')
    
    
    output = tf.matmul(fc, weights['out'], name='OUT') + biases['out']
    return output



    
def train_neural_network(x):
    prediction = convolutional_neural_network(x)
#    tf.add_to_collection('losses',tf.reduce_mean( tf.nn.softmax_cross_entropy_with_logits(logits=prediction, labels=y )))
#    cost = tf.add_n(tf.get_collection('losses'))
    cost = tf.reduce_mean( tf.nn.softmax_cross_entropy_with_logits(logits=prediction, labels=y) )
    optimizer = tf.train.AdamOptimizer(learning_rate=learning_rate,beta1=0.9).minimize(cost)
    
    hm_epochs = 40
    with tf.Session() as sess:
        sess.run(tf.global_variables_initializer())
        
        writer =  tf.summary.FileWriter('./graphs_model2',sess.graph)
        
        correct = tf.equal(tf.argmax(prediction, 1), tf.argmax(y, 1))
        accuracy = tf.reduce_mean(tf.cast(correct, 'float'))
        saver = tf.train.Saver()
        #saver_w = tf.train.Saver(weights)
        #saver_b = tf.train.Saver(biases)
        global checkpoint
        
        
        if not(checkpoint==0):
            saver.restore(sess, file_path)
            book = load_workbook("model_2.xlsx")
            sheet1 = book.get_sheet_by_name("Sheet2")
        else :
            book = Workbook()
            sheet1 = book.create_sheet("Sheet2") #.add_sheet('Sheet1',cell_overwrite_ok=True)
            book.save('model_2.xlsx')
        
        
        for epoch in range(hm_epochs):
            # apply cross validation
            x_tr, y_tr, x_val, y_val = cross_validation(Xtr,Ytr,5,epoch)
            train_labels = make_one_hot(y_tr,n_classes)
            val_labels = make_one_hot(y_val,n_classes)
            # make data augmentaion operation 
            # simply flip images
            print ('augmented data : just horizontal flip for quarter of data')
            trainNum = np.shape(x_tr)[0]
            flip_random_mask = np.random.choice(range(trainNum),size=int(trainNum/4),replace=False)
            x_tr[flip_random_mask,:,:,:] = x_tr[flip_random_mask,:,::-1,:]
            # random shift right
            right_shift_mask = np.random.choice(range(trainNum),size=int(trainNum/4),replace=False)
            random_right_shift = np.random.randint(2,6)
            zero_pad = np.zeros([int(trainNum/4),image_height,random_right_shift,image_channels])
            x_tr[right_shift_mask,:,:,:] = np.concatenate((zero_pad,x_tr[right_shift_mask,:,:-random_right_shift,:]),axis=2)
            # random shift down
            down_shift_mask = np.random.choice(range(trainNum),size=int(trainNum/4),replace=False)
            random_down_shift = np.random.randint(2,6)
            zero_pad2 = np.zeros([int(trainNum/4),random_down_shift,image_width,image_channels])
            x_tr[down_shift_mask,:,:,:] = np.concatenate((zero_pad2,x_tr[down_shift_mask,:-random_down_shift,:,:]),axis=1)
            
            
            #print np.shape(x_tr) , np.shape(x_val) ," data shape"
            train = readCIFAR.DataSet(x_tr, train_labels)
            validation = readCIFAR.DataSet(x_val, val_labels)
            
#            dev_labels = make_one_hot(Ytr[:32000],n_classes)
#            dev = readCIFAR.DataSet(Xtr[:32000], dev_labels)
            epoch_loss = 0
            start_time = time.time()
            
            
            
            for _ in range(int(train.num_examples/batch_size)):
                epoch_x, epoch_y = train.next_batch(batch_size) # TODO: change to train.next_batch
                _, c = sess.run([optimizer, cost], feed_dict={x: epoch_x, y: epoch_y ,keep_prob: 0.5})
                epoch_loss += c
            
            duration = time.time() - start_time
            print("Epoch lasts for : ",duration)
            print('Epoch', epoch, 'completed out of',hm_epochs,'loss:',epoch_loss)
            # save a checkpoint
            saver.save(sess,file_path)
            saver_w.save(sess,weights_path)
            saver_b.save(sess,biases_path)
            
            checkpoint = 1
            print("weights have been saved")
            # test against the validation data set
            percent = 0
            for batch in range(int(validation.num_examples/batch_size)):
                val_x, val_y = validation.next_batch(batch_size)
                percent += accuracy.eval({x:val_x, y:val_y, keep_prob: 1})
                #print percent
            percent /= (validation.num_examples/batch_size)
            print('Validation Accuracy:', percent)
            sheet1.append({"A":percent})
            book.save('model_2.xlsx')    
            
            if(epoch%10==9 or epoch==hm_epochs-1):
                # check accuracy aginst test data
                percent = 0
                for batch in range(int(test.num_examples/1000)):
                    #correct = tf.equal(tf.argmax(prediction, 1), tf.argmax(y, 1))
                    #accuracy = tf.reduce_mean(tf.cast(correct, 'float'))
                    test_x, test_y = test.next_batch(1000) # TODO : some thing wrong
                    percent += accuracy.eval({x:test_x, y:test_y, keep_prob: 1})
                print('Accuracy:',percent/10)

# here changed the dropout probability = 0.75 at train and 1 at test & validation
# learning rate = 0.001
# layer1 conv -> num of kernels = 64
# layer2 conv -> num of kernels = 64
# layer3 fully connected -> 512 neurons
# last layer -> softmax for 10 classes
# read from checkpoint or not (make it 0 in first time run only )
# then put it equal to 1 always
checkpoint = 0
# please specify the file where to store the model
file_path = "/home/ti1080/GP_TEAM2_salama/model_2.ckpt"
weights_path = "/home/ti1080/GP_TEAM2_salama/model_2_weights.ckpt"
biases_path = "/home/ti1080/GP_TEAM2_salama/model_2_biases.ckpt"
train_neural_network(x)
